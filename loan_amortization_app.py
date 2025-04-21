import streamlit as st
import pandas as pd
from datetime import datetime
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import numbers, Font, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule

# --- Sheet layout constants ---
PARAM_ROWS = 5  # Number of parameter lines
HEADER_ROW = PARAM_ROWS + 1
DATA_START_ROW = HEADER_ROW + 1

# --- Column headers ---
HEADERS = [
    "Payment #", "Due Date", "Scheduled Payment", "Scheduled Interest", "Scheduled Principal",
    "Extra Payment Made", "Actual Payment Date", "Actual Payment Made",
    "Late Fee", "Total Payment Due", "Remaining Balance", "Underpayment Flag"
]

# Build mapping from header name to Excel column letter

def build_col_map(headers):
    return {name: get_column_letter(i + 1) for i, name in enumerate(headers)}

st.title("📊 Commercial Loan Amortization Schedule")

# --- User inputs ---
st.markdown(
    """
    Enter the loan details below to generate a customized amortization schedule.  
    All amounts are in USD and rates are annual percentages.
    """
)
col1, col2 = st.columns(2)
with col1:
    loan_amount = st.number_input(
        "Loan Amount ($)",
        min_value=0.0,
        step=10000.0,
        format="%.2f",
        help="Total principal amount of the loan."
    )
    term_years = st.number_input(
        "Loan Term (Years)",
        min_value=1,
        step=1,
        help="Number of years until final maturity."
    )
with col2:
    interest_rate = st.number_input(
        "Annual Interest Rate (%)",
        min_value=0.0,
        step=0.1,
        format="%.3f",
        help="Annual interest rate (e.g., 7.5 for 7.5%)."
    )
    amortization_years = st.number_input(
        "Amortization Period (Years)",
        min_value=1,
        step=1,
        help="Period over which loan is amortized."
    )
start_date = st.date_input(
    "Loan Start Date",
    value=datetime.today(),
    help="Date of the first payment or drawdown."
)

if st.button("Generate Amortization Schedule"):
    # Derived values
    monthly_rate = interest_rate / 100 / 12
    term_months = int(term_years * 12)
    amort_months = int(amortization_years * 12)

    # Calculate monthly payment
    if monthly_rate == 0:
        # Zero-interest: divide principal evenly
        monthly_payment = loan_amount / amort_months
    else:
        monthly_payment = (
            loan_amount * (monthly_rate * (1 + monthly_rate) ** amort_months)
            / ((1 + monthly_rate) ** amort_months - 1)
        )

    # Build schedule data
    balance = loan_amount
    schedule = []
    for m in range(1, term_months + 1):
        interest_amt = balance * monthly_rate
        principal_amt = monthly_payment - interest_amt
        balance -= principal_amt
        if balance < 0:
            principal_amt += balance
            balance = 0
        schedule.append({
            "Payment #": m,
            "Due Date": (start_date + pd.DateOffset(months=m)).date(),
            "Scheduled Payment": round(monthly_payment, 2),
            "Scheduled Interest": round(interest_amt, 2),
            "Scheduled Principal": round(principal_amt, 2),
        })
        if balance <= 0:
            break

    st.success("Schedule generated. Download your Excel tracker below.")

    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Loan Payment Tracker"

            # Write loan parameters at top with labels in column B and values in column C
    params = [
        (1, "Loan Amount:", loan_amount),
        (2, "Interest Rate (%):", interest_rate),
        (3, "Loan Term (Years):", term_years),
        (4, "Amortization Period (Years):", amortization_years),
        (5, "Loan Start Date:", start_date)
    ]
    date_fmt = "m/d/yyyy"
    for row_num, label, val in params:
        # Label cell in column B, right aligned
        label_cell = ws.cell(row=row_num, column=2, value=label)
        label_cell.alignment = Alignment(horizontal="right")
        # Value cell in column C, left aligned with indent
        val_cell = ws.cell(row=row_num, column=3, value=val)
        val_cell.alignment = Alignment(horizontal="left", indent=1)
        # Apply formatting: currency for loan amount, date format for start date
        if row_num == 1:
            val_cell.number_format = numbers.FORMAT_CURRENCY_USD_SIMPLE
        elif row_num == 5:
            val_cell.number_format = date_fmt

    # Write headers
    col_map = build_col_map(HEADERS)
    for idx, header in enumerate(HEADERS, start=1):
        ws.cell(row=HEADER_ROW, column=idx, value=header)
    # Style header row
    thick = Side(border_style="thick", color="000000")
    for idx in range(1, len(HEADERS) + 1):
        cell = ws.cell(row=HEADER_ROW, column=idx)
        cell.font = Font(bold=True)
        cell.border = Border(bottom=thick)

    # Populate rows
    for i, row in enumerate(schedule):
        r = DATA_START_ROW + i
        # Static values
        ws.cell(row=r, column=1, value=row["Payment #"])
        ws.cell(row=r, column=2, value=row["Due Date"])
        # Scheduled Payment (override on final term for balloon)
        sp_col = col_map["Scheduled Payment"]
        if i == term_months - 1:
            # Final term: include balloon (remaining principal + standard payment)
            prev_balance = balance
            standard = row["Scheduled Payment"]
            ws[f"{sp_col}{r}"].value = f"={standard}+{prev_balance}"
        else:
            ws.cell(row=r, column=3, value=row["Scheduled Payment"])
        # Scheduled Interest & Principal
        ws.cell(row=r, column=4, value=row["Scheduled Interest"])
        ws.cell(row=r, column=5, value=row["Scheduled Principal"])

        # Column letters
        cm = col_map
        ld = cm["Due Date"]
        lad = cm["Actual Payment Date"]
        le = cm["Extra Payment Made"]
        lsp = cm["Scheduled Principal"]
        lsi = cm["Scheduled Interest"]
        lap = cm["Actual Payment Made"]
        ll = cm["Late Fee"]
        lt = cm["Total Payment Due"]
        lb = cm["Remaining Balance"]
        lf = cm["Underpayment Flag"]

        # Late Fee
        ws[f"{ll}{r}"].value = f"=IF(AND(ISNUMBER({lad}{r}),ISNUMBER({ld}{r}),{lad}{r}>{ld}{r}+10),35,0)"
        # Total Payment Due
        ws[f"{lt}{r}"].value = f"={lsp}{r}+{lsi}{r}+{ll}{r}"
        # Remaining Balance
        prev_ref = str(loan_amount) if i == 0 else f"{lb}{r-1}"
        ws[f"{lb}{r}"].value = f'=IF(AND(ISNUMBER({lap}{r}),ISNUMBER({ll}{r})),{prev_ref} - MIN({lsp}{r}+{le}{r},MAX(0,{lap}{r}-{lsi}{r}-{ll}{r})), "")'
        # Underpayment Flag
        ws[f"{lf}{r}"].value = f"=IF(AND(ISNUMBER({lap}{r}),{lap}{r}<{lt}{r}),\"UNDERPAID\",\"\")"

    # Conditional formatting: make 'UNDERPAID' red
    flag_col = col_map["Underpayment Flag"]
    flag_range = f"{flag_col}{DATA_START_ROW}:{flag_col}{DATA_START_ROW+len(schedule)-1}"
    red_font_rule = CellIsRule(
        operator='equal',
        formula=["\"UNDERPAID\""],
        font=Font(color="00FF0000")
    )
    ws.conditional_formatting.add(flag_range, red_font_rule)

    # Conditional formatting: make 'Actual Payment Made' red when underpaid
    ap_col = col_map["Actual Payment Made"]
    tpd_col = col_map["Total Payment Due"]
    ap_range = f"{ap_col}{DATA_START_ROW}:{ap_col}{DATA_START_ROW+len(schedule)-1}"
    red_ap_rule = CellIsRule(
        operator='lessThan',
        formula=[f"{tpd_col}{DATA_START_ROW}"],
        font=Font(color="00FF0000")
    )
    ws.conditional_formatting.add(ap_range, red_ap_rule)

        # Format dates as m/d/yyyy
    date_fmt = "m/d/yyyy"
    for col in ["Due Date", "Actual Payment Date"]:
        for row_cells in ws[f"{col_map[col]}{DATA_START_ROW}:{col_map[col]}{DATA_START_ROW+len(schedule)-1}"]:
            for cell in row_cells:
                cell.number_format = date_fmt

        # Auto-fit columns
    for idx, header in enumerate(HEADERS, start=1):
        col_letter = get_column_letter(idx)
        width = max(len(header) + 2, 12)
        ws.column_dimensions[col_letter].width = width

    # Format currency columns to simple USD
    acc_fmt = numbers.FORMAT_CURRENCY_USD_SIMPLE
    currency_cols = [
        "Scheduled Payment", "Scheduled Interest", "Scheduled Principal", 
        "Extra Payment Made", "Actual Payment Made", "Late Fee", 
        "Total Payment Due", "Remaining Balance"
    ]
    for name in currency_cols:
        col = col_map[name]
        for row_cells in ws[f"{col}{DATA_START_ROW}:{col}{DATA_START_ROW+len(schedule)-1}"]:
            for cell in row_cells:
                cell.number_format = acc_fmt

        # Set Remaining Balance column width explicitly
    rb_col = col_map["Remaining Balance"]
    ws.column_dimensions[rb_col].width = 18.1

        # Download
    output = BytesIO()
    wb.save(output)
    st.download_button(
        label="Download Tracker as Excel",
        data=output.getvalue(),
        file_name="Loan_Payment_Tracker.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
